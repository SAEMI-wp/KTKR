import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from .models import AttendanceMonthly, AttendanceDaily
from .utils import get_or_create_monthly_structure
from calendar import monthrange
import calendar
import tkinter as tk
from tkinter import ttk

# 스타일 상수 정의
class ExcelStyles:
    # 색상 정의
    ACTIVE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # 흰색
    INACTIVE_FILL = PatternFill(start_color="E9F2F9", end_color="E9F2F9", fill_type="solid")  # 회색
    HEADER_FILL = PatternFill(start_color="87cefa", end_color="87cefa", fill_type="solid")  # 하늘색
    SUBTOTAL_FILL = PatternFill(start_color="81c147", end_color="81c147", fill_type="solid")  # 연두색
    ORANGE_FILL = PatternFill(start_color="f8ad85", end_color="f8ad85", fill_type="solid")  # 주황색
    YELLOW_FILL = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")  # 노랑색
    HOLIDAY_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 연한 노랑
    
    # 테두리 정의
    THIN = Side(style='thin', color='000000')
    DOUBLE = Side(style='double', color='000000')
    DOTTED = Side(style='dotted', color='000000')
    MEDIUM = Side(style='medium', color='000000')

    # 폰트 정의
    TITLE_FONT = Font(size=18, bold=True)
    HEADER_FONT = Font(size=12, bold=True)
    NORMAL_FONT = Font(size=10)
    SMALL_FONT = Font(size=8)
    
    # 정렬 정의
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True)
    LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 엑셀 생성 클래스
class ExcelReportGenerator:
    def __init__(self, employee, year, month):
        self.employee = employee
        self.year = year
        self.month = month
        self.workbook = None
        self.worksheet = None
        self.styles = ExcelStyles()
        
    def generate_report(self):
        """가동보고서 엑셀 파일을 생성합니다."""
        try:
            # 구조체 기반으로 월별 데이터 가져오기
            monthly_data = get_or_create_monthly_structure(
                employee=self.employee,
                year=str(self.year),
                month=str(self.month)
            )
            
            if not monthly_data:
                raise ValueError("該当月の情報が見つかりません。")
            
            # 엑셀 워크북 생성
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "稼働報告書"

            # 1단계: 기본 스타일 적용
            self._apply_base_styles()
            
            # 2단계: 데이터 입력
            self._create_header_data()
            self._create_monthly_info_data(monthly_data)
            self._create_daily_table_data(monthly_data.daily_list)
            
            # 3단계: 디자인 적용
            self._apply_header_design()
            self._apply_table_design()
            self._apply_column_widths()
            self._apply_row_heights()
            # 표 테두리만 적용
            self._add_table_borders()
            
            # 4단계: 시트 보호 설정
            self._apply_sheet_protection()
            
            # 5단계: 세로 중앙 정렬 적용
            self._apply_vertical_center()
            
            return self.workbook
            
        except Exception as e:
            print(f"Excel generation error: {e}")
            raise

    def _apply_base_styles(self):
        """기본 스타일을 적용합니다."""
        # 시트 전체에 기본 비활성 스타일(회색 배경) 적용
        for row in self.worksheet.iter_rows(min_row=1, max_row=46, min_col=1, max_col=16):
            for cell in row:
                cell.fill = self.styles.INACTIVE_FILL

    def _create_header_data(self):
        """헤더 데이터를 입력합니다."""
        # 버전넘버
        self.worksheet['O1'] = "TA2025v1. 00"
        
        # 제목
        self.worksheet['B2'] = "稼 働 報 告 書"
        self.worksheet.merge_cells('B2:G3')
        
        # 보고서 시기
        self.worksheet['J2'] = f"{self.year}年 {self.month}月"
        self.worksheet.merge_cells('J2:K2')
        
        # 도장 섹션
        self.worksheet['M2'] = "承認"
        self.worksheet.merge_cells('M2:M3')
        
        self.worksheet['N2'] = "確認"
        self.worksheet.merge_cells('N2:N3')
        
        self.worksheet['O2'] = "申請"
        self.worksheet.merge_cells('O2:O3')
        
        self.worksheet.merge_cells('M4:M7')
        self.worksheet.merge_cells('N4:N7')
        self.worksheet.merge_cells('O4:O7')
        
        # 회사마크
        self.worksheet['B5'] = "(株)TEchAve"
        self.worksheet.merge_cells('B5:F5')
        
        # 작성자 정보
        self.worksheet['H7'] = "作成者 : "
        self.worksheet['I7'] = f"{self.employee.display_name or self.employee.employee_no}"
        self.worksheet.merge_cells('I7:K7')

    def _create_monthly_info_data(self, monthly_data):
        """월별 정보 데이터를 입력합니다."""
        # 상단 정보 (6-7행)
        self.worksheet['C6'] = "カレンダー : "
        self.worksheet['E6'] = monthly_data.base_calendar
        
        self.worksheet['H6'] = "PJ名 : "
        self.worksheet.merge_cells('I6:K6')
        self.worksheet['I6'] = monthly_data.project_name
        
        # 휴게 시간과 기준 시간 (7행)
        self.worksheet['C7'] = "昼休み区分 : "
        self.worksheet['E7'] = f"{monthly_data.break_minutes}分間"
        self.worksheet['F7'] = "基準時間 : "
        self.worksheet['G7'] = f"{monthly_data.standard_work_hours}Hr"
        
        # 하단 통계 정보 (44-45행)
        # 라벨 (44행)
        labels = ["出勤日", "年次\n有給", "特別\n有給", "無給日", "常勤", "控除", "残業", "深夜", "休日", "休日\n深夜", "残業換算(h)"]
        columns = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
        
        # 제목
        self.worksheet['B44'] = "報告値"
        self.worksheet.merge_cells('B44:C45')
        self.worksheet['B44'].alignment = Alignment(horizontal="center")
        self.worksheet.merge_cells('N44:O44')

        for i, (label, col) in enumerate(zip(labels, columns)):
            cell = self.worksheet[f'{col}44']
            cell.value = label
        
        # 데이터 (45행)
        self.worksheet['D45'] = f"{monthly_data.work_days:.1f}"
        self.worksheet['E45'] = f"{monthly_data.paid_leave_days:.1f}"
        self.worksheet['F45'] = f"{getattr(monthly_data, 'special_paid_leave_days', 0):.1f}"
        self.worksheet['G45'] = f"{getattr(monthly_data, 'unpaid_leave_days', 0):.1f}"
        self.worksheet['H45'] = f"{monthly_data.total_regular_work_hours:.2f}"
        self.worksheet['I45'] = f"{monthly_data.total_deduction_hours:.2f}"
        self.worksheet['J45'] = f"{monthly_data.total_overtime_hours:.2f}"
        self.worksheet['K45'] = f"{monthly_data.total_late_night_overtime_hours:.2f}"
        self.worksheet['L45'] = f"{getattr(monthly_data, 'holiday_overtime_hours', 0):.2f}"
        self.worksheet['M45'] = f"{getattr(monthly_data, 'holiday_late_night_overtime_hours', 0):.2f}"
        overtime_conversion = f"{getattr(monthly_data, 'overtime_conversion_hours', monthly_data.total_overtime_hours + monthly_data.total_late_night_overtime_hours):.2f}"
        self.worksheet.merge_cells('N45:O45')
        self.worksheet['N45'] = overtime_conversion

    def _create_daily_table_data(self, daily_list):
        """일별 정보 테이블 데이터를 입력합니다."""
        # 테이블 헤더 (11행) - 개행문자 포함
        headers = [
            "月/日", 
            "曜\n日", 
            "勤務区分", 
            "代休/振替\nの勤務日", 
            "作業開始\n時 刻", 
            "作業終了\n時 刻", 
            "常 勤\n(Hr)", 
            "控 除\n(Hr)", 
            "残 業\n(Hr)", 
            "深 夜\n(Hr)", 
            "小 計\n(Hr)", 
            "実施作業内容・備考"
        ]
        
        for col, header in enumerate(headers, 2):  # B열부터 시작
            cell = self.worksheet.cell(row=11, column=col)
            cell.value = header
        self.worksheet.merge_cells('M11:O11')

        # 해당 월의 총 일수 계산
        _, last_day = monthrange(self.year, self.month)
        
        # daily_list를 날짜별로 딕셔너리로 변환 (빠른 검색을 위해)
        daily_dict = {}
        for daily in daily_list:
            daily_dict[daily.date.day] = daily
        
        # 테이블 데이터 입력
        current_row = 12
        for day in range(1, last_day + 1):
            if current_row > 42:  # 42행을 넘어가면 중단
                break
                
            # 날짜 생성
            current_date = date(self.year, self.month, day)
            
            # B열: 날짜 (6/1 형태)
            self.worksheet[f'B{current_row}'] = f"{self.month}/{day}"
            
            # C열: 요일
            weekday_names = ["月", "火", "水", "木", "金", "土", "日"]
            weekday = weekday_names[current_date.weekday()]
            self.worksheet[f'C{current_row}'] = weekday
            
            # D열부터 O열까지: 해당 날짜의 데이터가 있으면 입력
            if day in daily_dict:
                daily = daily_dict[day]
                
                # D열: 근무구분
                work_type = daily.work_type or "-"
                if work_type == "出勤":
                    self.worksheet[f'D{current_row}'] = ""
                else:
                    self.worksheet[f'D{current_row}'] = work_type
                
                # E열: 대휴/대체 근무일
                if daily.alternative_work_date:
                    self.worksheet[f'E{current_row}'] = daily.alternative_work_date.strftime("%m/%d")
                else:
                    self.worksheet[f'E{current_row}'] = " "
                
                # F열: 작업 시작 시간
                if daily.start_time:
                    self.worksheet[f'F{current_row}'] = daily.start_time.strftime("%H:%M")
                else:
                    self.worksheet[f'F{current_row}'] = " "
                
                # G열: 작업 종료 시간
                if daily.end_time:
                    self.worksheet[f'G{current_row}'] = daily.end_time.strftime("%H:%M")
                else:
                    self.worksheet[f'G{current_row}'] = " "
                
                # H열: 상근 시간
                if daily.regular_work_hours is not None:
                    self.worksheet[f'H{current_row}'] = f"{daily.regular_work_hours:.2f}"
                else:
                    self.worksheet[f'H{current_row}'] = " "
                
                # I열: 공제 시간
                if daily.deduction_hours is not None:
                    self.worksheet[f'I{current_row}'] = f"{daily.deduction_hours:.2f}"
                else:
                    self.worksheet[f'I{current_row}'] = " "
                
                # J열: 잔업 시간
                if daily.overtime_hours is not None:
                    self.worksheet[f'J{current_row}'] = f"{daily.overtime_hours:.1f}"
                else:
                    self.worksheet[f'J{current_row}'] = " "
                
                # K열: 심야 시간
                if daily.late_night_overtime_hours is not None:
                    self.worksheet[f'K{current_row}'] = f"{daily.late_night_overtime_hours:.1f}"
                else:
                    self.worksheet[f'K{current_row}'] = " "
                
                # L열: 소계 시간 (상근 + 잔업 + 심야 - 공제)
                regular = daily.regular_work_hours or 0
                overtime = daily.overtime_hours or 0
                late_night = daily.late_night_overtime_hours or 0
                deduction = daily.deduction_hours or 0
                subtotal = regular + overtime + late_night - deduction
                if subtotal > 0:
                    self.worksheet[f'L{current_row}'] = f"{subtotal:.2f}"
                else:
                    self.worksheet[f'L{current_row}'] = " "
                
                # M열: 실시 작업 내용・비고
                self.worksheet[f'M{current_row}'] = daily.notes or " "
            
            else:
                # 데이터가 없는 경우 토요일/일요일 자동 휴일 설정
                if weekday == "土":  # 토요일
                    self.worksheet[f'D{current_row}'] = "休日"
                elif weekday == "日":  # 일요일
                    self.worksheet[f'D{current_row}'] = "休日(法)"
                else:
                    self.worksheet[f'D{current_row}'] = " "
                
                # 나머지 열은 빈 값으로 설정
                for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                    self.worksheet[f'{col}{current_row}'] = " "
            
            current_row += 1

        # 합계 표시
        self.worksheet['B43'] = "合 計"
        self.worksheet.merge_cells('B43:C43')
        
        # 합계 계산 (H, I, J, K, L열)
        sum_columns = ['H', 'I', 'J', 'K', 'L']
        for col in sum_columns:
            self.worksheet[f'{col}43'] = f"=SUM({col}12:{col}42)"

    def _apply_header_design(self):
        """헤더 디자인을 적용합니다."""
        # 버전넘버 스타일
        self.worksheet['O1'].font = self.styles.SMALL_FONT
        self.worksheet['O1'].alignment = self.styles.RIGHT_ALIGN
        
        # 제목 스타일
        self.worksheet['B2'].font = self.styles.TITLE_FONT
        
        # 보고서 시기 스타일
        self.worksheet['J2'].font = self.styles.HEADER_FONT
        
        # 도장 섹션 스타일
        for col in ['M', 'N', 'O']:
            cell = self.worksheet[f'{col}2']
            cell.font = Font(size=9)
            cell.alignment = self.styles.CENTER_ALIGN
            cell.border = Border(top=self.styles.THIN, left=self.styles.THIN, right=self.styles.THIN)
            
            bottom_cell = self.worksheet[f'{col}3']
            bottom_cell.border = Border(left=self.styles.THIN, right=self.styles.THIN)
        
        # 도장 섹션 하단 테두리
        for col in ['M', 'N', 'O']:
            for row in range(4, 8):
                cell = self.worksheet[f'{col}{row}']
                if row == 4:
                    cell.border = Border(top=self.styles.THIN, left=self.styles.THIN, right=self.styles.THIN)
                elif row == 7:
                    cell.border = Border(left=self.styles.THIN, right=self.styles.THIN, bottom=self.styles.THIN)
                else:
                    cell.border = Border(left=self.styles.THIN, right=self.styles.THIN)
        
        # 회사마크 스타일
        self.worksheet['B5'].font = Font(size=11, bold=True)
        
        # 작성자 정보 스타일
        self.worksheet['H7'].font = self.styles.NORMAL_FONT
        self.worksheet['I7'].font = self.styles.NORMAL_FONT
        self.worksheet['H7'].border = Border(bottom=self.styles.DOTTED)
        self.worksheet['I7'].border = Border(bottom=self.styles.DOTTED)
        self.worksheet['J7'].border = Border(bottom=self.styles.DOTTED)
        self.worksheet['K7'].border = Border(bottom=self.styles.DOTTED)
        
        # 월별 정보 스타일
        self.worksheet['H6'].border = Border(bottom=self.styles.DOTTED)
        self.worksheet['I6'].border = Border(bottom=self.styles.DOTTED)
        self.worksheet['J6'].border = Border(bottom=self.styles.DOTTED)
        self.worksheet['K6'].border = Border(bottom=self.styles.DOTTED)

    def _apply_table_design(self):
        """테이블 디자인을 적용합니다."""
        # 전체를 흰색으로 변경
        for row in self.worksheet.iter_rows(min_row=1, max_row=46, min_col=1, max_col=16):
            for cell in row:
                cell.fill = self.styles.ACTIVE_FILL
        
        # 헤더 배경색 적용 (B11:O11)
        for col in range(2, 16):
            self.worksheet.cell(row=11, column=col).fill = self.styles.HEADER_FILL
            self.worksheet.cell(row=11, column=col).alignment = Alignment(horizontal="center", wrap_text=True)
        
        # 소계, 합계 셀 강조색 적용 (연두)
        for row in range(12, 43):
            self.worksheet[f'L{row}'].fill = self.styles.SUBTOTAL_FILL
        for col in ['H', 'I', 'J', 'K', 'L']:
            self.worksheet[f'{col}43'].fill = self.styles.SUBTOTAL_FILL
        
        # 하단 통계 정보 배경색
        for col in range(4, 16):  # D(4)~O(15)
            self.worksheet.cell(row=44, column=col).fill = self.styles.ORANGE_FILL
            self.worksheet.cell(row=45, column=col).fill = self.styles.YELLOW_FILL
        
        # 테이블 스타일링 (M열 병합)
        for row in range(12, 43):
            self.worksheet.merge_cells(f'M{row}:O{row}')
        
        # 44행 라벨 중앙 정렬 (horizontal만 center)
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            self.worksheet[f'{col}44'].alignment = Alignment(horizontal="center", wrap_text=True)
        
        # 45행 데이터 중앙 정렬 (horizontal만 center)
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            self.worksheet[f'{col}45'].alignment = Alignment(horizontal="center")
        
        # 합계 행 스타일
        self.worksheet['B43'].alignment = Alignment(horizontal="center")
        for col in ['H', 'I', 'J', 'K', 'L']:
            self.worksheet[f'{col}43'].alignment = Alignment(horizontal="right")
        
        # 휴일 관련 스타일링 적용
        for row in range(12, 43):
            work_type_cell = self.worksheet[f'D{row}']
            work_type_value = work_type_cell.value
            
            # 휴일 관련 키워드 체크
            holiday_keywords = ["休日", "休日(法)", "振替(休)", "振替(法)", "祝日"]
            is_holiday = any(keyword in str(work_type_value) for keyword in holiday_keywords)
            
            if is_holiday:
                # C열 요일 셀 스타일링: 연한 노랑 배경, 빨간색 글씨
                weekday_cell = self.worksheet[f'C{row}']
                weekday_cell.fill = self.styles.HOLIDAY_FILL
                weekday_cell.font = Font(color="FF0000")
        # 날짜/요일 등은 horizontal만 center로 지정
        for row in range(12, 43):
            self.worksheet[f'B{row}'].alignment = Alignment(horizontal="center")  # 날짜
            self.worksheet[f'C{row}'].alignment = Alignment(horizontal="center")  # 요일
        # 숫자열 오른쪽 정렬
        for row in range(12, 43):
            for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L']:
                self.worksheet[f'{col}{row}'].alignment = Alignment(horizontal="right")
        # 비고(실시작업내용)는 왼쪽 정렬
        for row in range(12, 43):
            for col in ['M', 'N', 'O']:
                self.worksheet[f'{col}{row}'].alignment = Alignment(horizontal="left")

    def _apply_column_widths(self):
        """헤더 기준으로 열 너비를 고정하고, 일부 열은 수동 조정."""
        headers = [
            "月/日", "曜\n日", "勤務区分", "代休/振替\nの勤務日", "作業開始\n時 刻", "作業終了\n時 刻",
            "常 勤\n(Hr)", "控 除\n(Hr)", "残 業\n(Hr)", "深 夜\n(Hr)", "小 計\n(Hr)", "実施作業内容・備考"
        ]
        # A열, P열(1): 최소화
        self.worksheet.column_dimensions['A'].width = 2.5
        self.worksheet.column_dimensions['P'].width = 2.5

        # B~L열
        for idx, header in enumerate(headers[:-1], 2):  # B~L
            column_letter = get_column_letter(idx)
            max_line = max(header.split('\n'), key=len)
            length = len(max_line)
            if any(ord(char) > 127 for char in max_line):
                length = int(length * 1.7)
            # F열(작업시작시각, idx==6): width=6
            if column_letter == 'F':
                self.worksheet.column_dimensions[column_letter].width = 6
            else:
                self.worksheet.column_dimensions[column_letter].width = length + 1.5  # 여유 최소화

        # M,N,O열 (実施作業内容・備考): 여유 없이 3등분
        last_header = headers[-1]
        max_line = max(last_header.split('\n'), key=len)
        length = len(max_line)
        if any(ord(char) > 127 for char in max_line):
            length = int(length * 1.7)
        total_width = length + 4  # 여유 최소
        each_width = total_width / 3
        for col in ['M', 'N', 'O']:
            self.worksheet.column_dimensions[col].width = each_width

        # 열 너비 상세 조정
        self.worksheet.column_dimensions['C'].width += 1.5
        self.worksheet.column_dimensions['E'].width += 1.8
        self.worksheet.column_dimensions['F'].width += 3
        self.worksheet.column_dimensions['D'].width += 1.8
        self.worksheet.column_dimensions['G'].width += 1.8

    def _apply_row_heights(self):
        """행 높이 일괄 조정 + 특정 행만 별도 조정"""
        # 1. 전체 기본값
        for row in range(1, 47):
            self.worksheet.row_dimensions[row].height = 18
        # 2. 11행(헤더)만 넉넉하게
        self.worksheet.row_dimensions[11].height = 32
        # 3. 특정 행만 별도 조정 (덮어쓰기)
        row_heights = {
            1: 10,
            2: 18,
            3: 18,
            4: 6,
            8: 10,
            9: 10,
            10: 10,
            43: 15,
            46: 10
        }
        for row, height in row_heights.items():
            self.worksheet.row_dimensions[row].height = height

    def _apply_sheet_protection(self):
        """시트 보호를 설정합니다."""
        self.worksheet.protection.sheet = True
        for row in range(1, 47):
            for col in range(1, 17):
                self.worksheet.cell(row=row, column=col).protection = Protection(locked=False)
        self.worksheet.row_dimensions[43].height = 18
        self.worksheet.row_dimensions[44].height = 30

    def _add_table_borders(self, start_row=11, end_row=42):
        """테이블 테두리를 openpyxl 권장 방식에 따라 명확하고 단계적으로 추가합니다."""
        
        # 각 셀에 적용할 테두리 정보를 미리 계산하여 저장할 딕셔너리
        borders = {}

        # 1. 모든 셀의 기본 테두리 계산 (내부선: 세로 Thin, 가로 Dotted)
        for r in range(start_row, end_row + 1):
            for c in range(2, 16):
                top = self.styles.DOTTED if r > start_row else None
                left = self.styles.THIN if c > 2 else None
                borders[(r, c)] = Border(top=top, left=left)

        # 2. 외곽선 및 특별선으로 정보 업데이트 (덮어쓰기)
        for r in range(start_row, end_row + 1):
            for c in range(2, 16):
                current_border = borders[(r,c)]
                
                # 왼쪽/오른쪽 외곽선 (Thin)
                if c == 8:
                    left = self.styles.MEDIUM
                else:
                    left = self.styles.THIN if c == 2 else current_border.left

                if c == 12:
                    right = self.styles.MEDIUM
                else:
                    right = self.styles.THIN if c == 15 else current_border.right

                # 위쪽 외곽선 (Thin)
                if r == start_row:
                    # 헤더 행: H-L열은 MEDIUM, 나머지는 THIN
                    if 7 < c < 13:
                        top = self.styles.MEDIUM
                    else:
                        top = self.styles.THIN
                else:
                    # 내부 행: Step 1에서 계산된 Dotted 스타일을 유지
                    top = current_border.top
                
                # 아래쪽 외곽선 (Thin) 및 헤더 이중선
                if r == start_row: # 헤더 행
                    bottom = self.styles.DOUBLE
                elif r == end_row: # 마지막 행
                    if c > 7 and c < 13:
                        bottom = self.styles.MEDIUM
                    else:
                        bottom = self.styles.THIN
                else:
                    bottom = current_border.bottom
                
                borders[(r, c)] = Border(left=left, right=right, top=top, bottom=bottom)

        # 3. 계산된 테두리 정보를 실제 셀에 한 번에 적용
        for (r, c), border_style in borders.items():
            self.worksheet.cell(row=r, column=c).border = border_style
        
        # 4. 서브 테이블 테두리 설정
        self.worksheet.cell(row=43, column=2).border = Border(bottom=self.styles.MEDIUM, top=self.styles.MEDIUM, left=self.styles.MEDIUM)
        self.worksheet.cell(row=43, column=3).border = Border(bottom=self.styles.MEDIUM, top=self.styles.MEDIUM, right=self.styles.MEDIUM)

        for col in range(8, 13):
            if col == 8:
                left = self.styles.MEDIUM
            else:
                left = self.styles.THIN

            if col == 12:
                right = self.styles.MEDIUM
            else:
                right = self.styles.THIN
            
            top = self.styles.MEDIUM
            bottom = self.styles.MEDIUM
            self.worksheet.cell(row=43, column=col).border = Border(left=left, right=right, top=top, bottom=bottom)
        
        for col in range(2, 16):
            left = Side(border_style="thin", color="000000") # self.styles.THIN 대신 직접 Side 객체를 사용
            right = Side(border_style="thin", color="000000") # self.styles.THIN
            top = Side(border_style="medium", color="000000") # self.styles.MEDIUM (row=44는 항상 top이 MEDIUM)
            bottom = Side(border_style="thin", color="000000") # self.styles.THIN (기본값)

            if col == 2:
                left = Side(border_style="medium", color="000000") # self.styles.MEDIUM
            elif col == 3:    
                right = Side(border_style="medium", color="000000") # self.styles.MEDIUM
            elif col == 8:
                left = Side(border_style="medium", color="000000") # self.styles.MEDIUM
            elif col == 14:
                left = Side(border_style="medium", color="000000") # self.styles.MEDIUM
            elif col == 15:
                right = Side(border_style="medium", color="000000") # self.styles.MEDIUM
        
            self.worksheet.cell(row=44, column=col).border = Border(left=left, right=right, top=top, bottom=bottom)

        for col in range(2, 16):
            left = Side(border_style="thin", color="000000") # self.styles.THIN
            right = Side(border_style="thin", color="000000") # self.styles.THIN
            top = Side(border_style="thin", color="000000") 
            bottom = Side(border_style="medium", color="000000") # self.styles.MEDIUM (row=45는 항상 bottom이 MEDIUM)

            if col == 2:
                left = Side(border_style="medium", color="000000") # self.styles.MEDIUM
            elif col == 3:
                right = Side(border_style="medium", color="000000") # self.styles.MEDIUM
            elif col == 8:
                left = Side(border_style="medium", color="000000") # self.styles.MEDIUM
            elif col == 14:
                left = Side(border_style="medium", color="000000") # self.styles.MEDIUM
            elif col == 15:
                right = Side(border_style="medium", color="000000") # self.styles.MEDIUM

            self.worksheet.cell(row=45, column=col).border = Border(left=left, right=right, top=top, bottom=bottom)

    def _apply_vertical_center(self):
        """세로 중앙 정렬을 적용합니다."""
        for row in self.worksheet.iter_rows(min_row=1, max_row=46, min_col=1, max_col=16):
            for cell in row:
                # 기존 정렬이 있으면 horizontal은 유지, vertical만 center로
                if cell.alignment:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical="center",
                        wrap_text=cell.alignment.wrap_text
                    )
                else:
                    cell.alignment = Alignment(vertical="center") 